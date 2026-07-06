"""
Import trades and occupation codes from GLOBAL_TRADE_BANK_COMPLETE.xlsx
into complete_trade_bank.json and trade_occupation_codes.json.

Preserves duties/responsibilities for trades that already exist in the bank.
New trades receive a generic duty template.
"""
from __future__ import annotations

import json
import re
import shutil
import sys
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

DATA_DIR = ROOT / "data"
BANK_PATH = DATA_DIR / "complete_trade_bank.json"
CODES_PATH = DATA_DIR / "trade_occupation_codes.json"
XLSX_DEFAULT = Path(r"z:\Projects\GLOBAL_TRADE_BANK_COMPLETE.xlsx")
XLSX_PROJECT = DATA_DIR / "GLOBAL_TRADE_BANK_COMPLETE.xlsx"

SUB_CATEGORY_TO_INDUSTRY: dict[str, str] = {
    "Access & Safety": "Construction & Infrastructure",
    "Accommodation": "Hospitality & Tourism",
    "Accounting": "Professional & Business Services",
    "Accounts": "Professional & Business Services",
    "Allied Health": "Healthcare & Medical",
    "Analysis": "Professional & Business Services",
    "Aquaculture & Fishing": "Agriculture & Primary",
    "Architecture": "Professional & Business Services",
    "Audio": "IT & Technology",
    "Audit": "Professional & Business Services",
    "Automation": "IT & Technology",
    "Bookkeeping": "Professional & Business Services",
    "Brand": "Retail & Consumer",
    "Broadcasting": "IT & Technology",
    "Building Management": "Construction & Infrastructure",
    "Chemical": "Manufacturing & Industrial",
    "Childcare": "Education & Training",
    "Civil & Earthworks": "Construction & Infrastructure",
    "Civil & Structural": "Construction & Infrastructure",
    "Cleaning": "Professional & Business Services",
    "Clinical & Nursing": "Healthcare & Medical",
    "Cold Chain": "Logistics & Supply Chain",
    "Community": "Aged Care & Disability",
    "Compensation": "Professional & Business Services",
    "Compliance": "Professional & Business Services",
    "Content": "IT & Technology",
    "Contracts": "Professional & Business Services",
    "Creative": "IT & Technology",
    "Credit & Risk": "Professional & Business Services",
    "Dairy & Livestock": "Agriculture & Primary",
    "Data & Analytics": "IT & Technology",
    "Defence": "Security & Safety",
    "Design": "IT & Technology",
    "Development": "IT & Technology",
    "Digital Marketing": "Retail & Consumer",
    "Digital Media": "IT & Technology",
    "Earth Sciences": "Energy & Resources",
    "Electrical": "Construction & Infrastructure",
    "Electrical & Low Voltage": "Construction & Infrastructure",
    "Environmental": "Energy & Resources",
    "Events & Catering": "Hospitality & Tourism",
    "Fabrication & Welding": "Construction & Infrastructure",
    "Facilities": "Construction & Infrastructure",
    "Farming": "Agriculture & Primary",
    "Fashion": "Retail & Consumer",
    "Financial Services": "Professional & Business Services",
    "Finishing Trades": "Construction & Infrastructure",
    "Fitness": "Retail & Consumer",
    "Flight Crew": "Hospitality & Tourism",
    "Food & Beverage": "Hospitality & Tourism",
    "Food Manufacturing": "Manufacturing & Industrial",
    "Food Science": "Manufacturing & Industrial",
    "Forestry": "Agriculture & Primary",
    "General Manufacturing": "Manufacturing & Industrial",
    "Geotechnical": "Construction & Infrastructure",
    "Government": "Professional & Business Services",
    "Ground Operations": "Logistics & Supply Chain",
    "Grounds": "Agriculture & Primary",
    "HR Business Partner": "Professional & Business Services",
    "HR Generalist": "Professional & Business Services",
    "HR Management": "Professional & Business Services",
    "Hair & Beauty": "Retail & Consumer",
    "Horticulture": "Agriculture & Primary",
    "Household": "Aged Care & Disability",
    "Industrial": "Manufacturing & Industrial",
    "Industrial Design": "Manufacturing & Industrial",
    "Industrial Relations": "Professional & Business Services",
    "Infrastructure & Networks": "IT & Technology",
    "Insulation & Fire Protection": "Construction & Infrastructure",
    "Insurance": "Professional & Business Services",
    "Interior Design": "Professional & Business Services",
    "Jewellery": "Retail & Consumer",
    "Laboratory": "Healthcare & Medical",
    "Landscaping & External Works": "Construction & Infrastructure",
    "Legal Practice": "Professional & Business Services",
    "Life Sciences": "Healthcare & Medical",
    "Maritime Crew": "Logistics & Supply Chain",
    "Maritime Officers": "Logistics & Supply Chain",
    "Marketing Management": "Retail & Consumer",
    "Mathematics": "Professional & Business Services",
    "Mechanical": "Construction & Infrastructure",
    "Mechanical Services": "Construction & Infrastructure",
    "Media Buying": "Retail & Consumer",
    "Media Production": "IT & Technology",
    "Medical Support": "Healthcare & Medical",
    "Mental Health": "Healthcare & Medical",
    "Mining": "Energy & Resources",
    "Oil & Gas": "Energy & Resources",
    "PR": "Retail & Consumer",
    "Payroll": "Professional & Business Services",
    "Pest Control": "Security & Safety",
    "Pharmaceutical": "Manufacturing & Industrial",
    "Pharmacy": "Healthcare & Medical",
    "Physical Sciences": "Energy & Resources",
    "Port & Logistics": "Logistics & Supply Chain",
    "Print Media": "Manufacturing & Industrial",
    "Process Engineering": "Manufacturing & Industrial",
    "Product Design": "IT & Technology",
    "Property Management": "Professional & Business Services",
    "Recreation": "Hospitality & Tourism",
    "Recruitment": "Professional & Business Services",
    "Religious": "Aged Care & Disability",
    "Renewable Energy": "Energy & Resources",
    "Research": "Professional & Business Services",
    "Retail Operations": "Retail & Consumer",
    "Robotics": "IT & Technology",
    "Roofing & Cladding": "Construction & Infrastructure",
    "Safety": "Security & Safety",
    "Sales": "Retail & Consumer",
    "Security Services": "Security & Safety",
    "Shipping & Maritime": "Logistics & Supply Chain",
    "Social Work": "Aged Care & Disability",
    "Software Development": "IT & Technology",
    "Specialty Law": "Professional & Business Services",
    "Specialty Retail": "Retail & Consumer",
    "Sports": "Retail & Consumer",
    "Structural & Framing": "Construction & Infrastructure",
    "Supervision & Management": "Construction & Infrastructure",
    "Survey & Engineering Support": "Construction & Infrastructure",
    "Teaching": "Education & Training",
    "Telecommunications": "IT & Technology",
    "Therapy": "Healthcare & Medical",
    "Training": "Education & Training",
    "Training & Development": "Education & Training",
    "Transport & Freight": "Logistics & Supply Chain",
    "Transport Operations": "Logistics & Supply Chain",
    "Treasury": "Professional & Business Services",
    "Valuation": "Professional & Business Services",
    "Vehicle Trades": "Manufacturing & Industrial",
    "Veterinary": "Healthcare & Medical",
    "Warehousing": "Logistics & Supply Chain",
    "Waste Management": "Manufacturing & Industrial",
}

INDUSTRY_META = {
    "Construction & Infrastructure": {"icon": "🏗️", "color": "#1A3C5E"},
    "Healthcare & Medical": {"icon": "🏥", "color": "#C0392B"},
    "Agriculture & Primary": {"icon": "🌾", "color": "#27AE60"},
    "Hospitality & Tourism": {"icon": "🍽️", "color": "#E67E22"},
    "Logistics & Supply Chain": {"icon": "🚛", "color": "#2980B9"},
    "Manufacturing & Industrial": {"icon": "🏭", "color": "#7F8C8D"},
    "Energy & Resources": {"icon": "⚡", "color": "#F39C12"},
    "IT & Technology": {"icon": "💻", "color": "#8E44AD"},
    "Aged Care & Disability": {"icon": "👴", "color": "#16A085"},
    "Education & Training": {"icon": "🏫", "color": "#D35400"},
    "Security & Safety": {"icon": "🔒", "color": "#2C3E50"},
    "Retail & Consumer": {"icon": "🛒", "color": "#E91E63"},
    "Professional & Business Services": {"icon": "💼", "color": "#34495E"},
}

CODE_COLUMNS = [
    ("ANZSCO", 3, 4),
    ("ISCO", 5, 6),
    ("NOC", 7, None),
    ("SOC_UK", 8, None),
    ("SOC", 9, None),
    ("UAE_MOL", 10, None),
    ("KSA_MOL", 11, None),
    ("QATAR_MOI", 12, None),
    ("KUWAIT_MOI", 13, None),
    ("OMAN_MOL", 14, None),
    ("BAHRAIN_LMRA", 15, None),
    ("SSOC", 16, None),
]


def _clean_code(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text or text.lower() in {"n/a", "na", "-", "—"}:
        return ""
    return text


def _generic_duties(trade_name: str, industry: str) -> tuple[list[str], list[str]]:
    responsibilities = [
        f"Perform {trade_name} duties in accordance with organisational policies, industry standards, and applicable legislation.",
        f"Maintain safe work practices and comply with health, safety, and environmental requirements relevant to {industry}.",
        "Collaborate with supervisors, colleagues, and stakeholders to deliver quality outcomes.",
        "Maintain accurate records, documentation, and reporting as required.",
        "Support continuous improvement and professional development within the role.",
    ]
    duties = [
        f"Carry out core {trade_name} tasks as directed by work orders, specifications, and supervisor instructions.",
        f"Apply technical knowledge and practical skills required for the {trade_name} role.",
        "Inspect own work for quality and compliance with standards before sign-off.",
        "Use tools, equipment, and systems safely and report defects or hazards promptly.",
        "Communicate progress, issues, and completion status to the supervisor or team lead.",
        "Follow site or workplace policies including PPE, access control, and emergency procedures.",
        "Participate in toolbox talks, briefings, and training as required.",
        "Maintain a clean, organised, and safe work area.",
    ]
    return responsibilities, duties


def _load_excel_rows(xlsx_path: Path) -> list[dict]:
    responsibilities = [
        f"Perform {trade_name} duties in accordance with organisational policies, industry standards, and applicable legislation.",
        f"Maintain safe work practices and comply with health, safety, and environmental requirements relevant to {industry}.",
        f"Collaborate with supervisors, colleagues, and stakeholders to deliver quality outcomes.",
        f"Maintain accurate records, documentation, and reporting as required.",
        f"Support continuous improvement and professional development within the role.",
    ]
    duties = [
        f"Carry out core {trade_name} tasks as directed by work orders, specifications, and supervisor instructions.",
        f"Apply technical knowledge and practical skills required for the {trade_name} role.",
        f"Inspect own work for quality and compliance with standards before sign-off.",
        f"Use tools, equipment, and systems safely and report defects or hazards promptly.",
        f"Communicate progress, issues, and completion status to the supervisor or team lead.",
        f"Follow site or workplace policies including PPE, access control, and emergency procedures.",
        f"Participate in toolbox talks, briefings, and training as required.",
        f"Maintain a clean, organised, and safe work area.",
    ]
    return responsibilities, duties


def _load_excel_rows(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["3 - ALL TRADES MASTER"]
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        trade_name = (row[2] or "").strip() if row[2] else ""
        sub_category = (row[1] or "").strip() if row[1] else "General"
        if not trade_name:
            continue
        industry = SUB_CATEGORY_TO_INDUSTRY.get(sub_category, "Professional & Business Services")
        codes = {}
        for system, code_idx, title_idx in CODE_COLUMNS:
            code = _clean_code(row[code_idx] if code_idx < len(row) else None)
            if not code:
                continue
            title = trade_name
            if title_idx is not None and title_idx < len(row):
                title = _clean_code(row[title_idx]) or trade_name
            codes[system] = {
                "code": code,
                "title": title,
                "confidence": "CONFIRMED",
                "note": "GLOBAL_TRADE_BANK_COMPLETE.xlsx",
            }
        rows.append(
            {
                "trade": trade_name,
                "sub_category": sub_category,
                "industry": industry,
                "occupation_codes": codes,
            }
        )
    wb.close()
    return rows


def _index_existing(bank: dict) -> dict[str, tuple[dict, dict, dict]]:
    """trade lower name -> (industry_obj, category_obj, trade_obj)"""
    index: dict[str, tuple[dict, dict, dict]] = {}
    for ind in bank.get("industries", []):
        for cat in ind.get("categories", []):
            for trade in cat.get("trades", []):
                index[trade["trade"].strip().lower()] = (ind, cat, trade)
    return index


def main() -> None:
    xlsx_path = XLSX_PROJECT if XLSX_PROJECT.exists() else XLSX_DEFAULT
    if not xlsx_path.exists():
        raise SystemExit(f"Excel file not found: {xlsx_path}")

    if not XLSX_PROJECT.exists() and xlsx_path != XLSX_PROJECT:
        shutil.copy2(xlsx_path, XLSX_PROJECT)
        print(f"Copied Excel to {XLSX_PROJECT}")

    excel_rows = _load_excel_rows(xlsx_path)
    with open(BANK_PATH, encoding="utf-8") as f:
        bank = json.load(f)

    existing_index = _index_existing(bank)
    updated_codes = 0
    preserved = 0
    added = 0

    # industry name -> industry object
    industries: dict[str, dict] = {}
    for ind in bank.get("industries", []):
        industries[ind["industry"]] = ind

    for ind_name, meta in INDUSTRY_META.items():
        if ind_name not in industries:
            industries[ind_name] = {
                "industry": ind_name,
                "icon": meta["icon"],
                "color": meta["color"],
                "categories": [],
            }

    # category key -> category object
    cat_index: dict[tuple[str, str], dict] = {}
    for ind in industries.values():
        for cat in ind["categories"]:
            cat_index[(ind["industry"], cat["category"])] = cat

    for item in excel_rows:
        trade_name = item["trade"]
        key = trade_name.strip().lower()
        codes = item["occupation_codes"]
        anzsco = codes.get("ANZSCO", {})

        if key in existing_index:
            ind, cat, trade = existing_index[key]
            trade["occupation_codes"] = codes
            if anzsco.get("code"):
                trade["anzsco_code"] = anzsco["code"]
            if anzsco.get("title"):
                trade["anzsco_title"] = anzsco["title"]
            updated_codes += 1
            preserved += 1
            continue

        industry_name = item["industry"]
        sub_category = item["sub_category"]
        ind = industries[industry_name]
        cat_key = (industry_name, sub_category)
        if cat_key not in cat_index:
            cat_obj = {"category": sub_category, "trades": []}
            ind["categories"].append(cat_obj)
            cat_index[cat_key] = cat_obj
        else:
            cat_obj = cat_index[cat_key]

        responsibilities, duties = _generic_duties(trade_name, industry_name)
        new_trade = {
            "trade": trade_name,
            "anzsco_code": anzsco.get("code", ""),
            "anzsco_title": anzsco.get("title", trade_name),
            "responsibilities": responsibilities,
            "duties": duties,
            "occupation_codes": codes,
        }
        cat_obj["trades"].append(new_trade)
        added += 1

    # Sort and rebuild industries list
    industry_order = list(INDUSTRY_META.keys())
    new_industries = []
    for name in industry_order:
        if name not in industries:
            continue
        ind = industries[name]
        ind["categories"].sort(key=lambda c: c["category"].lower())
        for cat in ind["categories"]:
            cat["trades"].sort(key=lambda t: t["trade"].lower())
        if ind["categories"]:
            new_industries.append(ind)

    total_trades = sum(len(c["trades"]) for i in new_industries for c in i["categories"])
    bank["industries"] = new_industries
    bank["meta"] = {
        **bank.get("meta", {}),
        "version": "4.0",
        "source": "GLOBAL_TRADE_BANK_COMPLETE.xlsx",
        "total_industries": len(new_industries),
        "total_trades": total_trades,
    }

    with open(BANK_PATH, "w", encoding="utf-8") as f:
        json.dump(bank, f, indent=2, ensure_ascii=False)
        f.write("\n")

    # trade_occupation_codes.json for runtime merge fallback
    codes_export = []
    for item in excel_rows:
        codes_export.append(
            {
                "trade_name": item["trade"],
                "codes": {
                    "ANZSCO": item["occupation_codes"].get("ANZSCO"),
                    "ISCO08": item["occupation_codes"].get("ISCO"),
                    "NOC": item["occupation_codes"].get("NOC"),
                    "SOC_UK": item["occupation_codes"].get("SOC_UK"),
                    **{
                        k: v
                        for k, v in item["occupation_codes"].items()
                        if k not in {"ANZSCO", "ISCO", "NOC", "SOC_UK"}
                    },
                },
            }
        )
    # strip None entries
    for entry in codes_export:
        entry["codes"] = {k: v for k, v in entry["codes"].items() if v}

    with open(CODES_PATH, "w", encoding="utf-8") as f:
        json.dump(codes_export, f, indent=2, ensure_ascii=False)
        f.write("\n")

    print(f"Excel trades processed: {len(excel_rows)}")
    print(f"Existing trades — codes updated: {updated_codes}, duties preserved: {preserved}")
    print(f"New trades added: {added}")
    print(f"Total industries: {len(new_industries)}, total trades: {total_trades}")
    print(f"Wrote {BANK_PATH}")
    print(f"Wrote {CODES_PATH}")


if __name__ == "__main__":
    main()
