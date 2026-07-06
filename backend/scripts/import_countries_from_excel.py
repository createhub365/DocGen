"""Import countries from GLOBAL_TRADE_BANK_COMPLETE.xlsx into countries.js."""
import json
import re
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

try:
    import pycountry
except ImportError:
    pycountry = None

ROOT = Path(__file__).resolve().parents[2]
EXCEL = ROOT / "backend" / "data" / "GLOBAL_TRADE_BANK_COMPLETE.xlsx"
CONTACT_CACHE = ROOT / "backend" / "data" / "countries_contact.json"
COUNTRIES_JS = ROOT / "frontend" / "src" / "data" / "countries.js"
HELPERS_MARKER = "// ── HELPER FUNCTIONS"

COUNTRY_SYSTEM_OVERRIDE = {
    "NZ": "ANZSCO",
    "AU": "ANZSCO",
    "GB": "SOC_UK",
    "IE": "SOC_UK",
    "CA": "NOC",
    "US": "SOC",
    "IN": "NCO",
    "SG": "SSOC",
    "AE": "UAE_MOL",
    "SA": "KSA_MOL",
    "QA": "QATAR_MOI",
    "KW": "KUWAIT_MOI",
    "OM": "OMAN_MOL",
    "BH": "BAHRAIN_LMRA",
    "JO": "ISCO",
}
DISPLAY_NAMES = {
    "US": "United States",
    "GB": "United Kingdom",
    "BA": "Bosnia & Herzegovina",
    "CI": "Ivory Coast",
    "KN": "Saint Kitts & Nevis",
    "VC": "Saint Vincent",
    "ST": "São Tomé & Príncipe",
    "AE": "United Arab Emirates",
    "CZ": "Czech Republic",
    "CD": "DR Congo",
    "CG": "Congo",
    "CV": "Cape Verde",
    "AG": "Antigua & Barbuda",
}

NAME_TO_CODE = {
    "United States of America": "US",
    "United Kingdom": "GB",
    "Saint Kitts and Nevis": "KN",
    "Saint Vincent and Grenadines": "VC",
    "Cote d'Ivoire": "CI",
    "Côte d'Ivoire": "CI",
    "Ivory Coast": "CI",
    "Czechia": "CZ",
    "Czech Republic": "CZ",
    "Bosnia and Herzegovina": "BA",
    "Bosnia & Herzegovina": "BA",
    "North Macedonia": "MK",
    "Macedonia": "MK",
    "Congo": "CG",
    "Republic of the Congo": "CG",
    "Congo (Brazzaville)": "CG",
    "DR Congo": "CD",
    "Democratic Republic of the Congo": "CD",
    "Congo (DRC)": "CD",
    "East Timor": "TL",
    "Timor-Leste": "TL",
    "Sao Tome and Principe": "ST",
    "São Tomé & Príncipe": "ST",
    "Eswatini": "SZ",
    "Swaziland": "SZ",
    "Myanmar": "MM",
    "Burma": "MM",
    "South Korea": "KR",
    "Korea, Republic of": "KR",
    "North Korea": "KP",
    "Korea, Democratic People's Republic of": "KP",
    "Russia": "RU",
    "Russian Federation": "RU",
    "Viet Nam": "VN",
    "Vietnam": "VN",
    "UAE": "AE",
    "United Arab Emirates": "AE",
    "Palestine": "PS",
    "State of Palestine": "PS",
    "Taiwan": "TW",
    "Hong Kong": "HK",
    "Macau": "MO",
    "Macao": "MO",
    "Vatican City": "VA",
    "Holy See": "VA",
    "Trinidad and Tobago": "TT",
    "Antigua and Barbuda": "AG",
    "Antigua & Barbuda": "AG",
    "Dominican Republic": "DO",
    "Papua New Guinea": "PG",
    "New Zealand": "NZ",
    "Sri Lanka": "LK",
    "South Africa": "ZA",
    "Saudi Arabia": "SA",
    "Costa Rica": "CR",
    "El Salvador": "SV",
    "Cape Verde": "CV",
    "Cabo Verde": "CV",
    "Equatorial Guinea": "GQ",
    "Central African Republic": "CF",
    "Sierra Leone": "SL",
    "Burkina Faso": "BF",
    "Guinea-Bissau": "GW",
    "South Sudan": "SS",
    "Turkey": "TR",
    "Türkiye": "TR",
    "Laos": "LA",
    "Lao People's Democratic Republic": "LA",
    "Brunei": "BN",
    "Brunei Darussalam": "BN",
    "Syria": "SY",
    "Syrian Arab Republic": "SY",
    "Iran": "IR",
    "Iran, Islamic Republic of": "IR",
    "Bolivia": "BO",
    "Bolivia, Plurinational State of": "BO",
    "Venezuela": "VE",
    "Venezuela, Bolivarian Republic of": "VE",
    "Tanzania": "TZ",
    "Tanzania, United Republic of": "TZ",
    "Moldova": "MD",
    "Moldova, Republic of": "MD",
    "Micronesia": "FM",
    "Micronesia, Federated States of": "FM",
    "Bahamas": "BS",
    "Bahamas, The": "BS",
    "Gambia": "GM",
    "Gambia, The": "GM",
}


def map_occupation_system(raw: str) -> str:
    s = (raw or "").strip().lower()
    if not s or s == "none":
        return "ISCO"
    if "anzsco" in s:
        return "ANZSCO"
    if "o*net" in s or "onet" in s:
        return "SOC"
    if s == "soc 2020":
        return "SOC_UK"
    if s.startswith("noc") and "2021" in s:
        return "NOC"
    if s == "nco 2015":
        return "NCO"
    if "mohre" in s:
        return "UAE_MOL"
    if "mhrsd" in s:
        return "KSA_MOL"
    if "moi" in s and "qchp" in s:
        return "QATAR_MOI"
    if "moi profession" in s:
        return "KUWAIT_MOI"
    if "moi" in s and "kuwait" not in s and "qatar" not in s:
        return "QATAR_MOI"
    if "mol profession" in s and "oman" not in s:
        return "OMAN_MOL"
    if "lmra" in s:
        return "BAHRAIN_LMRA"
    if s == "ssoc 2020" or s.startswith("ssoc"):
        return "SSOC"
    return "ISCO"


def flag_emoji(code: str) -> str:
    if not code or len(code) != 2:
        return "🌐"
    return "".join(chr(0x1F1E6 + ord(c) - ord("A")) for c in code.upper())


def load_contact_meta() -> dict[str, dict]:
    if CONTACT_CACHE.exists():
        return json.loads(CONTACT_CACHE.read_text(encoding="utf-8"))

    url = "https://restcountries.com/v3.1/all?fields=cca2,name,idd,currencies"
    req = urllib.request.Request(url, headers={"User-Agent": "DocGenPro/1.0"})
    with urllib.request.urlopen(req, timeout=60) as resp:
        payload = json.loads(resp.read().decode("utf-8"))

    meta: dict[str, dict] = {}
    for item in payload:
        code = item.get("cca2")
        if not code:
            continue
        root = (item.get("idd") or {}).get("root") or ""
        suffixes = (item.get("idd") or {}).get("suffixes") or [""]
        phone = f"{root}{suffixes[0]}" if root else ""
        currencies = item.get("currencies") or {}
        currency = next(iter(currencies.keys()), "")
        names = item.get("name") or {}
        meta[code] = {
            "phone": phone,
            "currency": currency,
            "commonName": names.get("common") or code,
        }

    CONTACT_CACHE.write_text(json.dumps(meta, indent=2), encoding="utf-8")
    return meta


def resolve_country_code(excel_name: str) -> str | None:
    if excel_name in NAME_TO_CODE:
        return NAME_TO_CODE[excel_name]
    if pycountry:
        try:
            return pycountry.countries.search_fuzzy(excel_name)[0].alpha_2
        except LookupError:
            pass
    return None


def load_excel_rows():
    wb = load_workbook(EXCEL, read_only=True, data_only=True)
    ws = wb["1 - COUNTRIES"]
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        name = row[1]
        if not name or "TOTAL" in str(name):
            continue
        rows.append({
            "name": str(name).strip(),
            "region": (row[2] or "").strip(),
            "subRegion": (row[3] or "").strip(),
            "codeSystem": (row[4] or "").strip(),
            "basedOn": (row[5] or "").strip(),
            "unMember": (row[6] or "").strip(),
            "notes": (row[7] or "").strip() if row[7] else "",
        })
    wb.close()
    return rows


def esc(s: str) -> str:
    return s.replace("\\", "\\\\").replace('"', '\\"')


def country_line(c: dict) -> str:
    parts = [
        f'code: "{c["code"]}"',
        f'name: "{esc(c["name"])}"',
        f'flag: "{c["flag"]}"',
        f'phone: "{esc(c["phone"])}"',
        f'currency: "{esc(c["currency"])}"',
        f'region: "{esc(c["region"])}"',
        f'subRegion: "{esc(c.get("subRegion", ""))}"',
        f'codeSystem: "{esc(c.get("codeSystem", ""))}"',
        f'occupationSystem: "{esc(c.get("occupationSystem", "ISCO"))}"',
        f'basedOn: "{esc(c.get("basedOn", ""))}"',
    ]
    if c.get("notes"):
        parts.append(f'notes: "{esc(c["notes"])}"')
    return "  { " + ", ".join(parts) + " },"


def load_helpers_template() -> str:
    default = """
export const getCountryByCode = (code) =>
  COUNTRIES.find((c) => c.code === code)

const COUNTRY_ALIASES = {
  UAE: 'AE',
  UK: 'GB',
}

export const getCountryByName = (name) => {
  if (!name) return undefined
  const alias = COUNTRY_ALIASES[name]
  if (alias) return getCountryByCode(alias)
  return COUNTRIES.find(
    (c) => c.name === name || c.name.toLowerCase() === name.toLowerCase()
  )
}

export const getCountryCode = (nameOrCode) => {
  if (!nameOrCode) return undefined
  if (nameOrCode.length === 2) {
    const byCode = getCountryByCode(nameOrCode.toUpperCase())
    if (byCode) return byCode.code
  }
  return getCountryByName(nameOrCode)?.code
}

export const getCountryFlag = (nameOrCode) => {
  const c = getCountryByCode(nameOrCode) || getCountryByName(nameOrCode)
  return c?.flag || '🌐'
}

export const getCountriesByRegion = (region) =>
  COUNTRIES.filter((c) => c.region === region)

export const getRegions = () =>
  [...new Set(COUNTRIES.map((c) => c.region))]

export const searchCountries = (query) => {
  const q = query.toLowerCase()
  return COUNTRIES.filter(
    (c) =>
      c.name.toLowerCase().includes(q) ||
      c.code.toLowerCase().includes(q) ||
      c.phone.includes(q)
  )
}

export const countryToOption = (country) => ({
  value: country.code,
  label: country.name,
  country,
})

export const COUNTRY_OPTIONS = COUNTRIES.map(countryToOption)

export const COUNTRY_OPTIONS_GROUPED = getRegions().map((region) => ({
  label: region,
  options: COUNTRIES.filter((c) => c.region === region).map(countryToOption),
}))

export const PRIORITY_COUNTRIES = [
  'NZ', 'AU', 'GB', 'CA', 'AE',
  'JO', 'SA', 'QA', 'KW', 'IN',
  'PH', 'LK', 'NP', 'BD',
]

export const PRIORITY_COUNTRY_OPTIONS = [
  {
    label: '⭐ Frequently Used',
    options: PRIORITY_COUNTRIES.map((code) => COUNTRIES.find((c) => c.code === code))
      .filter(Boolean)
      .map(countryToOption),
  },
  {
    label: '── All Countries ──',
    options: COUNTRIES.filter((c) => !PRIORITY_COUNTRIES.includes(c.code)).map(
      countryToOption
    ),
  },
]

export const getRegField = (countryCode) => {
  const fields = {
    NZ: { label: 'NZBN', placeholder: '9429041234567 (13 digits)', maxLength: 13, regLabel: 'NZBN' },
    AU: { label: 'ABN', placeholder: 'XX XXX XXX XXX (11 digits)', maxLength: 14, regLabel: 'ABN' },
    GB: { label: 'Companies House No.', placeholder: 'XXXXXXXX (8 digits)', maxLength: 8, regLabel: 'CRN' },
    CA: { label: 'CRA Business No.', placeholder: 'XXXXXXXXX (9 digits)', maxLength: 9, regLabel: 'BN' },
    AE: { label: 'TRN', placeholder: '100XXXXXXXXXX (15 digits)', maxLength: 15, regLabel: 'TRN' },
    SA: { label: 'Commercial Registration No.', placeholder: '10-digit CR number', maxLength: 10, regLabel: 'CR' },
    QA: { label: 'Commercial Registration No.', placeholder: 'QA CR number', maxLength: 15, regLabel: 'CR' },
    KW: { label: 'Commercial Registration No.', placeholder: 'Kuwait CR number', maxLength: 15, regLabel: 'CR' },
    JO: { label: 'Company Registration No.', placeholder: 'Jordan company number', maxLength: 15, regLabel: 'CR' },
    IN: { label: 'CIN / GSTIN', placeholder: '21-char CIN or 15-char GSTIN', maxLength: 21, regLabel: 'CIN' },
    US: { label: 'EIN (Tax ID)', placeholder: 'XX-XXXXXXX', maxLength: 10, regLabel: 'EIN' },
  }
  return (
    fields[countryCode] || {
      label: 'Registration Number',
      placeholder: 'Company registration number',
      maxLength: 50,
      regLabel: 'Registration No.',
    }
  )
}

export const getOccupationSystem = (countryCode) =>
  getCountryByCode(countryCode)?.occupationSystem || 'ISCO'

export const IMMIGRATION_COUNTRY_CODES = new Set(['NZ', 'AU', 'AE', 'JO'])

export const countryMatchesList = (countryCode, countryNames) => {
  if (!countryCode || !countryNames?.length) return true
  const meta = getCountryByCode(countryCode)
  if (!meta) return false
  const variants = new Set([
    meta.code,
    meta.name,
    meta.name.toLowerCase(),
    ...(countryCode === 'GB' ? ['UK', 'United Kingdom'] : []),
    ...(countryCode === 'AE' ? ['UAE', 'United Arab Emirates'] : []),
  ])
  return countryNames.some((name) => {
    if (!name) return false
    if (variants.has(name) || variants.has(name.toLowerCase())) return true
    return getCountryByName(name)?.code === countryCode
  })
}

export const resolveCatalogCountryName = (code, catalogCountries = []) => {
  const meta = getCountryByCode(code)
  if (!meta) return null
  const match = catalogCountries.find((c) => countryMatchesList(code, [c]))
  return match || meta.name
}

export default COUNTRIES
"""
    if not COUNTRIES_JS.exists():
        return default
    text = COUNTRIES_JS.read_text(encoding="utf-8")
    if HELPERS_MARKER in text:
        return text.split(HELPERS_MARKER, 1)[1]
    return default


def main():
    contact = load_contact_meta()
    helpers = load_helpers_template()
    excel_rows = load_excel_rows()
    merged: dict[str, dict] = {}
    unmatched: list[str] = []

    for row in excel_rows:
        code = resolve_country_code(row["name"])
        if not code:
            unmatched.append(row["name"])
            continue
        info = contact.get(code, {})
        merged[code] = {
            "code": code,
            "name": DISPLAY_NAMES.get(code) or info.get("commonName") or row["name"],
            "flag": flag_emoji(code),
            "phone": info.get("phone", ""),
            "currency": info.get("currency", ""),
            "region": row["region"],
            "subRegion": row["subRegion"],
            "codeSystem": row["codeSystem"],
            "occupationSystem": COUNTRY_SYSTEM_OVERRIDE.get(code)
            or map_occupation_system(row["codeSystem"]),
            "basedOn": row["basedOn"],
            "notes": row["notes"],
        }

    ordered = sorted(merged.values(), key=lambda c: (c["region"], c["name"]))
    lines = [
        "// World countries — imported from GLOBAL_TRADE_BANK_COMPLETE.xlsx (Sheet 1 - COUNTRIES)",
        "",
        "const COUNTRIES = [",
    ]
    lines.extend(country_line(c) for c in ordered)
    lines.append("]")
    lines.append("")
    lines.append(HELPERS_MARKER + helpers)

    COUNTRIES_JS.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {len(ordered)} countries to {COUNTRIES_JS}")
    if unmatched:
        print(f"Unmatched ({len(unmatched)}): {', '.join(unmatched[:10])}")


if __name__ == "__main__":
    main()
